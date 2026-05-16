import csv
import json
from io import StringIO, BytesIO
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from ideas.models import Idea
from feedbacks.models import Feedback
from comments.models import Comment


def _load_idea(idea_id, user):
    idea = get_object_or_404(Idea, id=idea_id)
    if idea.owner != user and user.role != 'admin':
        return None, None, None, Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)
    feedbacks = Feedback.objects.filter(idea=idea).select_related('reviewer', 'reviewer__userprofile').order_by('-created_at')
    comments  = Comment.objects.filter(idea=idea, is_deleted=False, parent__isnull=True).select_related('author').prefetch_related('replies__author').order_by('created_at')
    return idea, feedbacks, comments, None


def _level(f):
    return getattr(getattr(f.reviewer, 'userprofile', None), 'level', 'Bronze')


def _avg(fb_list):
    return round(sum(f.weighted_score for f in fb_list) / len(fb_list), 2) if fb_list else 0


def _safe(text):
    """Escape XML special chars for ReportLab Paragraph."""
    if not text:
        return ''
    return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def _filename(idea_id, title, ext):
    safe = title[:30].replace(' ', '_').replace('/', '_')
    return f'idea_{idea_id}_{safe}.{ext}'


# ── CSV ───────────────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_csv(request, idea_id):
    idea, feedbacks, comments, err = _load_idea(idea_id, request.user)
    if err:
        return err

    from datetime import datetime

    fb_list = list(feedbacks)
    comment_list = list(comments)
    replies_count = sum(c.replies.filter(is_deleted=False).count() for c in comment_list)
    comments_total = len(comment_list) + replies_count
    avg_score = _avg(fb_list)
    generated_at = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')

    avg_market = round(sum(f.market_score for f in fb_list) / len(fb_list), 1) if fb_list else 0
    avg_innovation = round(sum(f.innovation_score for f in fb_list) / len(fb_list), 1) if fb_list else 0
    avg_feasibility = round(sum(f.feasibility_score for f in fb_list) / len(fb_list), 1) if fb_list else 0
    avg_roi = round(sum(f.roi_score for f in fb_list) / len(fb_list), 1) if fb_list else 0

    out = StringIO()
    out.write('sep=,\r\n')
    w = csv.writer(out)

    def blank():
        w.writerow([])

    def section(title):
        blank()
        w.writerow([title.upper()])

    w.writerow(['IDEALAB IDEA REPORT'])
    w.writerow(['Generated At', generated_at])
    w.writerow(['Report Type', 'Structured CSV Export'])

    section('Executive Summary')
    w.writerow(['Metric', 'Value', 'Notes'])
    w.writerow(['Idea ID', idea.id, ''])
    w.writerow(['Title', idea.title, ''])
    w.writerow(['Owner', idea.owner.username, ''])
    w.writerow(['Sector', idea.sector, ''])
    w.writerow(['Status', idea.status.capitalize(), 'Current workflow status'])
    w.writerow(['Created', idea.created_at.strftime('%Y-%m-%d %H:%M'), 'UTC'])
    w.writerow(['Average SGV', avg_score, 'Average reviewer weighted score out of 25'])
    w.writerow(['Reviews', len(fb_list), 'Total submitted feedback entries'])
    w.writerow(['Comments', comments_total, 'Top-level comments plus replies'])

    section('Idea Narrative')
    w.writerow(['Area', 'Content'])
    w.writerow(['Description', idea.description])
    w.writerow(['Problem', idea.problem])
    w.writerow(['Solution', idea.solution])
    w.writerow(['Target Audience', idea.target])

    section('Score Breakdown')
    w.writerow(['Dimension', 'Average Score', 'Maximum Score', 'Completion %'])
    for label, value in [
        ('Market', avg_market),
        ('Innovation', avg_innovation),
        ('Feasibility', avg_feasibility),
        ('ROI', avg_roi),
    ]:
        completion = f'{round((value / 25) * 100, 1)}%' if value else '0%'
        w.writerow([label, value, 25, completion])

    section('Reviewer Feedback')
    w.writerow([
        'Review #', 'Reviewer', 'Reviewer Level', 'Date', 'Market /25',
        'Innovation /25', 'Feasibility /25', 'ROI /25', 'Raw /100',
        'Weighted /25', 'Most Helpful', 'Comment',
    ])
    if fb_list:
        for index, f in enumerate(fb_list, start=1):
            w.writerow([
                index,
                f.reviewer.username,
                _level(f),
                f.created_at.strftime('%Y-%m-%d %H:%M'),
                f.market_score,
                f.innovation_score,
                f.feasibility_score,
                f.roi_score,
                f.raw_score,
                round(f.weighted_score or 0, 2),
                'Yes' if f.is_helpful else 'No',
                f.comment,
            ])
    else:
        w.writerow(['-', 'No reviews submitted yet', '', '', '', '', '', '', '', '', '', ''])

    section('Community Discussion')
    w.writerow(['Item #', 'Type', 'Author', 'In Reply To', 'Date', 'Content'])
    if comment_list:
        row_number = 1
        for c in comment_list:
            w.writerow([
                row_number,
                'Comment',
                c.author.username,
                '',
                c.created_at.strftime('%Y-%m-%d %H:%M'),
                c.content,
            ])
            parent_number = row_number
            row_number += 1
            for r in c.replies.filter(is_deleted=False).order_by('created_at'):
                w.writerow([
                    row_number,
                    'Reply',
                    r.author.username,
                    f'Comment {parent_number}',
                    r.created_at.strftime('%Y-%m-%d %H:%M'),
                    r.content,
                ])
                row_number += 1
    else:
        w.writerow(['-', 'No comments yet', '', '', '', ''])

    section('Export Notes')
    w.writerow(['Note'])
    w.writerow(['Scores use the IdeaLab review scale: each dimension is scored out of 25 and the weighted SGV score is shown out of 25.'])
    w.writerow(['Open this file in Excel, Google Sheets, or LibreOffice and use Freeze Top Row or filters on the table sections if needed.'])

    resp = HttpResponse(('\ufeff' + out.getvalue()).encode('utf-8'), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{_filename(idea_id, idea.title, "csv")}"'
    return resp


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_xlsx(request, idea_id):
    idea, feedbacks, comments, err = _load_idea(idea_id, request.user)
    if err:
        return err

    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    fb_list = list(feedbacks)
    comment_list = list(comments)
    replies_count = sum(c.replies.filter(is_deleted=False).count() for c in comment_list)
    comments_total = len(comment_list) + replies_count

    avg_score = _avg(fb_list)
    avg_market = round(sum(f.market_score for f in fb_list) / len(fb_list), 1) if fb_list else 0
    avg_innovation = round(sum(f.innovation_score for f in fb_list) / len(fb_list), 1) if fb_list else 0
    avg_feasibility = round(sum(f.feasibility_score for f in fb_list) / len(fb_list), 1) if fb_list else 0
    avg_roi = round(sum(f.roi_score for f in fb_list) / len(fb_list), 1) if fb_list else 0

    primary = '681A15'
    primary_soft = 'F5F0EF'
    header_fill = 'EFE4E2'
    border_color = 'D8C7C4'
    muted = '6F5F5D'
    white = 'FFFFFF'
    green = '397A4E'

    thin = Side(style='thin', color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill('solid', fgColor=primary)
    section_fill = PatternFill('solid', fgColor=header_fill)
    soft_fill = PatternFill('solid', fgColor=primary_soft)
    white_fill = PatternFill('solid', fgColor=white)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Idea Report'

    def set_widths(sheet, widths):
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    def style_range(sheet, start_row, end_row, start_col, end_col, fill=None):
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if fill:
                    cell.fill = fill

    def title(sheet, text, columns):
        row = 1 if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None else sheet.max_row + 1
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
        cell = sheet.cell(row=row, column=1, value=text)
        cell.fill = title_fill
        cell.font = Font(color=white, bold=True, size=16)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[row].height = 28
        style_range(sheet, row, row, 1, columns, title_fill)
        return row

    def section(sheet, text, columns):
        row = sheet.max_row + 2
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
        cell = sheet.cell(row=row, column=1, value=text.upper())
        cell.fill = section_fill
        cell.font = Font(color=primary, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        sheet.row_dimensions[row].height = 22
        style_range(sheet, row, row, 1, columns, section_fill)
        return row

    def write_table(sheet, headers, rows, start_row=None):
        if start_row is None:
            start_row = sheet.max_row + 1
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color=primary)
            cell.fill = soft_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r_index, row_data in enumerate(rows, start=start_row + 1):
            for c_index, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=r_index, column=c_index, value=value)
                cell.border = border
                cell.fill = white_fill
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        end_row = start_row + len(rows)
        if rows:
            sheet.auto_filter.ref = f'A{start_row}:{get_column_letter(len(headers))}{end_row}'
        return end_row

    set_widths(ws, [22, 42, 22, 26, 20, 20, 20, 20])
    title(ws, 'IdeaLab Idea Report', 8)
    ws.append(['Generated At', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'), 'Report Type', 'Designed Excel Workbook'])
    style_range(ws, ws.max_row, ws.max_row, 1, 4, soft_fill)

    section(ws, 'Executive Summary', 8)
    summary_rows = [
        ['Idea ID', idea.id, 'Title', idea.title],
        ['Owner', idea.owner.username, 'Sector', idea.sector],
        ['Status', idea.status.capitalize(), 'Created', idea.created_at.strftime('%Y-%m-%d %H:%M')],
        ['Average SGV /25', avg_score, 'Reviews', len(fb_list)],
        ['Discussion Items', comments_total, 'Global Score', idea.global_score],
    ]
    write_table(ws, ['Metric', 'Value', 'Metric', 'Value'], summary_rows)

    section(ws, 'Idea Narrative', 8)
    narrative_rows = [
        ['Description', idea.description],
        ['Problem', idea.problem],
        ['Solution', idea.solution],
        ['Target Audience', idea.target],
    ]
    write_table(ws, ['Area', 'Content'], narrative_rows)

    section(ws, 'Score Breakdown', 8)
    score_rows = []
    for label, value in [
        ('Market', avg_market),
        ('Innovation', avg_innovation),
        ('Feasibility', avg_feasibility),
        ('ROI', avg_roi),
    ]:
        score_rows.append([label, value, 25, round((value / 25) * 100, 1) if value else 0])
    write_table(ws, ['Dimension', 'Average Score', 'Maximum Score', 'Completion %'], score_rows)

    feedback_ws = wb.create_sheet('Reviews')
    set_widths(feedback_ws, [12, 22, 18, 18, 14, 16, 16, 14, 14, 14, 16, 70])
    title(feedback_ws, 'Reviewer Feedback', 12)
    feedback_rows = []
    if fb_list:
        for index, f in enumerate(fb_list, start=1):
            feedback_rows.append([
                index,
                f.reviewer.username,
                _level(f),
                f.created_at.strftime('%Y-%m-%d %H:%M'),
                f.market_score,
                f.innovation_score,
                f.feasibility_score,
                f.roi_score,
                f.raw_score,
                round(f.weighted_score or 0, 2),
                'Yes' if f.is_helpful else 'No',
                f.comment,
            ])
    else:
        feedback_rows.append(['-', 'No reviews submitted yet', '', '', '', '', '', '', '', '', '', ''])
    write_table(
        feedback_ws,
        ['#', 'Reviewer', 'Level', 'Date', 'Market /25', 'Innovation /25', 'Feasibility /25', 'ROI /25', 'Raw /100', 'Weighted /25', 'Most Helpful', 'Comment'],
        feedback_rows,
    )

    discussion_ws = wb.create_sheet('Discussion')
    set_widths(discussion_ws, [10, 14, 22, 18, 18, 80])
    title(discussion_ws, 'Community Discussion', 6)
    discussion_rows = []
    if comment_list:
        row_number = 1
        for c in comment_list:
            discussion_rows.append([row_number, 'Comment', c.author.username, '', c.created_at.strftime('%Y-%m-%d %H:%M'), c.content])
            parent_number = row_number
            row_number += 1
            for r in c.replies.filter(is_deleted=False).order_by('created_at'):
                discussion_rows.append([row_number, 'Reply', r.author.username, f'Comment {parent_number}', r.created_at.strftime('%Y-%m-%d %H:%M'), r.content])
                row_number += 1
    else:
        discussion_rows.append(['-', 'No comments yet', '', '', '', ''])
    write_table(discussion_ws, ['#', 'Type', 'Author', 'In Reply To', 'Date', 'Content'], discussion_rows)

    for sheet in wb.worksheets:
        sheet.freeze_panes = 'A4'
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or 'left',
                    vertical='top',
                    wrap_text=True,
                )
        for row_number in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row_number].height = max(sheet.row_dimensions[row_number].height or 18, 18)
        for cells in sheet.iter_cols():
            for cell in cells:
                if cell.value == 'Yes':
                    cell.font = Font(color=green, bold=True)
                elif cell.value == 'No':
                    cell.font = Font(color=muted)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{_filename(idea_id, idea.title, "xlsx")}"'
    return resp


# ── JSON ──────────────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_json(request, idea_id):
    idea, feedbacks, comments, err = _load_idea(idea_id, request.user)
    if err:
        return err

    import datetime
    fb_list = list(feedbacks)

    def comment_node(c):
        return {
            'id': str(c.id), 'author': c.author.username,
            'content': c.content, 'date': c.created_at.isoformat(),
            'replies': [
                {'id': str(r.id), 'author': r.author.username,
                 'content': r.content, 'date': r.created_at.isoformat()}
                for r in c.replies.filter(is_deleted=False).order_by('created_at')
            ],
        }

    data = {
        'exported_at': datetime.datetime.utcnow().isoformat() + 'Z',
        'idea': {
            'id': idea.id, 'title': idea.title, 'owner': idea.owner.username,
            'sector': idea.sector, 'status': idea.status,
            'description': idea.description, 'problem': idea.problem,
            'solution': idea.solution, 'target': idea.target,
            'global_score': idea.global_score, 'created_at': idea.created_at.isoformat(),
        },
        'statistics': {
            'total_feedbacks': len(fb_list),
            'total_comments':  comments.count(),
            'avg_sgv_score':   _avg(fb_list),
            'avg_market':      round(sum(f.market_score      for f in fb_list) / len(fb_list), 1) if fb_list else 0,
            'avg_innovation':  round(sum(f.innovation_score  for f in fb_list) / len(fb_list), 1) if fb_list else 0,
            'avg_feasibility': round(sum(f.feasibility_score for f in fb_list) / len(fb_list), 1) if fb_list else 0,
            'avg_roi':         round(sum(f.roi_score         for f in fb_list) / len(fb_list), 1) if fb_list else 0,
        },
        'feedbacks': [
            {
                'reviewer': f.reviewer.username, 'level': _level(f),
                'market_score': f.market_score, 'innovation_score': f.innovation_score,
                'feasibility_score': f.feasibility_score, 'roi_score': f.roi_score,
                'raw_score': f.raw_score, 'weighted_score': round(f.weighted_score, 2),
                'is_helpful': f.is_helpful, 'comment': f.comment,
                'date': f.created_at.isoformat(),
            }
            for f in fb_list
        ],
        'comments': [comment_node(c) for c in comments],
    }

    resp = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json')
    resp['Content-Disposition'] = f'attachment; filename="{_filename(idea_id, idea.title, "json")}"'
    return resp


# ── PDF ───────────────────────────────────────────────────────────────────────

def _export_pdf_legacy(request, idea_id):
    idea, feedbacks, comments, err = _load_idea(idea_id, request.user)
    if err:
        return err

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    PRIMARY  = colors.HexColor('#681a15')
    LIGHT_BG = colors.HexColor('#f5f0ef')
    BORDER   = colors.HexColor('#e8e0df')
    GRAY     = colors.HexColor('#888888')
    BLACK    = colors.HexColor('#1a1a1a')

    fb_list = list(feedbacks)
    avg     = _avg(fb_list)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm, topMargin=16*mm, bottomMargin=16*mm)
    W = doc.width

    def S(name, **kw):
        return ParagraphStyle(name, **kw)

    s_section  = S('sh', fontSize=9,  textColor=PRIMARY, spaceAfter=6, spaceBefore=10, fontName='Helvetica-Bold')
    s_label    = S('lb', fontSize=8,  textColor=GRAY,    spaceAfter=2, fontName='Helvetica-Bold')
    s_body     = S('bd', fontSize=10, textColor=BLACK,   spaceAfter=6, fontName='Helvetica', leading=15)
    s_small    = S('sm', fontSize=8,  textColor=GRAY,    fontName='Helvetica')
    s_italic   = S('it', fontSize=9,  textColor=colors.HexColor('#444'), fontName='Helvetica-Oblique', leading=13)

    story = []

    # ── Cover
    cover = Table(
        [
            [Paragraph('IDEALAB - IDEA REPORT', S('cl', fontSize=8, textColor=colors.white, fontName='Helvetica-Bold'))],
            [Paragraph(_safe(idea.title),        S('ct', fontSize=20, textColor=colors.white, fontName='Helvetica-Bold', leading=26))],
            [Paragraph(
                f'Owner: {_safe(idea.owner.username)}   |   Sector: {_safe(idea.sector)}   |   '
                f'Status: {idea.status.capitalize()}   |   Date: {idea.created_at.strftime("%B %d, %Y")}',
                S('cm', fontSize=9, textColor=colors.HexColor('#ffcccc'), fontName='Helvetica', leading=14)
            )],
            [Paragraph(f'SGV Score: {avg}', S('cs', fontSize=14, textColor=colors.white, fontName='Helvetica-Bold'))],
        ],
        colWidths=[W]
    )
    cover.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), PRIMARY),
        ('TOPPADDING',    (0,0), (0,0),   14),
        ('TOPPADDING',    (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-2), 6),
        ('BOTTOMPADDING', (0,-1),(-1,-1), 16),
        ('LEFTPADDING',   (0,0), (-1,-1), 20),
        ('RIGHTPADDING',  (0,0), (-1,-1), 20),
    ]))
    story += [cover, Spacer(1, 8*mm)]

    # ── Stats row
    stats = Table([[
        [Paragraph(str(avg),           S('n1', fontSize=20, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
         Paragraph('AVG SGV',          S('l1', fontSize=7,  textColor=GRAY,    fontName='Helvetica-Bold', alignment=TA_CENTER))],
        [Paragraph(str(len(fb_list)),  S('n2', fontSize=20, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
         Paragraph('FEEDBACKS',        S('l2', fontSize=7,  textColor=GRAY,    fontName='Helvetica-Bold', alignment=TA_CENTER))],
        [Paragraph(str(comments.count()), S('n3', fontSize=20, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
         Paragraph('COMMENTS',         S('l3', fontSize=7,  textColor=GRAY,    fontName='Helvetica-Bold', alignment=TA_CENTER))],
    ]], colWidths=[W/3]*3)
    stats.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), LIGHT_BG),
        ('TOPPADDING',    (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('LINEAFTER',     (0,0), (1,-1),  0.5, BORDER),
    ]))
    story += [stats, Spacer(1, 6*mm)]

    # ── Idea details
    story.append(HRFlowable(width='100%', thickness=1.5, color=PRIMARY, spaceAfter=4))
    story.append(Paragraph('IDEA DETAILS', s_section))
    for lbl, val in [('Description', idea.description), ('Problem', idea.problem),
                     ('Solution', idea.solution), ('Target Audience', idea.target)]:
        story.append(Paragraph(lbl.upper(), s_label))
        story.append(Paragraph(_safe(val) or '—', s_body))

    # ── Dimension averages
    if fb_list:
        story.append(HRFlowable(width='100%', thickness=1.5, color=PRIMARY, spaceAfter=4))
        story.append(Paragraph('DIMENSION AVERAGES', s_section))
        bar_w = W - 100
        for i, (dim_name, dim_val) in enumerate([
            ('Market',      round(sum(f.market_score      for f in fb_list) / len(fb_list), 1)),
            ('Innovation',  round(sum(f.innovation_score  for f in fb_list) / len(fb_list), 1)),
            ('Feasibility', round(sum(f.feasibility_score for f in fb_list) / len(fb_list), 1)),
            ('ROI',         round(sum(f.roi_score         for f in fb_list) / len(fb_list), 1)),
        ]):
            fill = max(4, int((dim_val / 25) * bar_w))
            bar_inner = Table([['']], colWidths=[fill], rowHeights=[8])
            bar_inner.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), PRIMARY)]))
            bar_outer = Table([[bar_inner]], colWidths=[bar_w], rowHeights=[8])
            bar_outer.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), LIGHT_BG)]))
            row = Table([[
                Paragraph(dim_name, S(f'dn{i}', fontSize=9, textColor=BLACK, fontName='Helvetica')),
                bar_outer,
                Paragraph(f'{dim_val}/25', S(f'dv{i}', fontSize=9, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
            ]], colWidths=[60, bar_w, 40])
            row.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('BOTTOMPADDING', (0,0), (-1,-1), 5)]))
            story.append(row)
        story.append(Spacer(1, 4*mm))

    # ── Feedbacks
    story.append(HRFlowable(width='100%', thickness=1.5, color=PRIMARY, spaceAfter=4))
    story.append(Paragraph(f'FEEDBACKS ({len(fb_list)})', s_section))
    if not fb_list:
        story.append(Paragraph('No feedbacks yet.', s_small))
    for f in fb_list:
        helpful_txt = '  [Most Helpful]' if f.is_helpful else ''
        uid = str(f.reviewer.username)
        header = Table([[
            Paragraph(_safe(f.reviewer.username),
                      S(f'fh{uid}', fontSize=10, textColor=BLACK, fontName='Helvetica-Bold')),
            Paragraph(f'{_level(f)}{helpful_txt}',
                      S(f'fl{uid}', fontSize=8, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER)),
            Paragraph(f.created_at.strftime('%b %d, %Y'),
                      S(f'fd{uid}', fontSize=8, textColor=GRAY, fontName='Helvetica', alignment=TA_RIGHT)),
            Paragraph(f'{round(f.weighted_score, 2)}/100',
                      S(f'fs{uid}', fontSize=11, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
        ]], colWidths=[W*0.35, W*0.25, W*0.2, W*0.2])
        header.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('BOTTOMPADDING', (0,0), (-1,-1), 6)]))

        scores = Table([[
            [Paragraph('MARKET',      S(f'sp1{uid}', fontSize=7, textColor=GRAY,    fontName='Helvetica-Bold', alignment=TA_CENTER)),
             Paragraph(f'{f.market_score}/25',      S(f'sv1{uid}', fontSize=11, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER))],
            [Paragraph('INNOVATION',  S(f'sp2{uid}', fontSize=7, textColor=GRAY,    fontName='Helvetica-Bold', alignment=TA_CENTER)),
             Paragraph(f'{f.innovation_score}/25',  S(f'sv2{uid}', fontSize=11, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER))],
            [Paragraph('FEASIBILITY', S(f'sp3{uid}', fontSize=7, textColor=GRAY,    fontName='Helvetica-Bold', alignment=TA_CENTER)),
             Paragraph(f'{f.feasibility_score}/25', S(f'sv3{uid}', fontSize=11, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER))],
            [Paragraph('ROI',         S(f'sp4{uid}', fontSize=7, textColor=GRAY,    fontName='Helvetica-Bold', alignment=TA_CENTER)),
             Paragraph(f'{f.roi_score}/25',         S(f'sv4{uid}', fontSize=11, textColor=PRIMARY, fontName='Helvetica-Bold', alignment=TA_CENTER))],
        ]], colWidths=[W/4]*4)
        scores.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,-1), LIGHT_BG),
            ('TOPPADDING',    (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LINEAFTER',     (0,0), (2,-1),  0.5, BORDER),
        ]))

        fb_block = Table([[header], [scores], [Paragraph(_safe(f.comment) or '—', s_italic)]], colWidths=[W])
        fb_block.setStyle(TableStyle([
            ('BOX',           (0,0), (-1,-1), 0.5, BORDER),
            ('TOPPADDING',    (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('LEFTPADDING',   (0,0), (-1,-1), 12),
            ('RIGHTPADDING',  (0,0), (-1,-1), 12),
        ]))
        story += [fb_block, Spacer(1, 4*mm)]

    # ── Comments
    story.append(HRFlowable(width='100%', thickness=1.5, color=PRIMARY, spaceAfter=4))
    story.append(Paragraph(f'COMMUNITY COMMENTS ({comments.count()})', s_section))
    if not comments.count():
        story.append(Paragraph('No comments yet.', s_small))
    for c in comments:
        c_block = Table([[
            [Paragraph(_safe(c.author.username),
                       S(f'ca{str(c.id)[:8]}', fontSize=9, textColor=PRIMARY, fontName='Helvetica-Bold')),
             Paragraph(_safe(c.content), s_body)],
            Paragraph(c.created_at.strftime('%b %d, %Y'), s_small),
        ]], colWidths=[W*0.82, W*0.18])
        c_block.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,-1), LIGHT_BG),
            ('TOPPADDING',    (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('LEFTPADDING',   (0,0), (-1,-1), 10),
            ('RIGHTPADDING',  (0,0), (-1,-1), 10),
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('LINEBEFORE',    (0,0), (0,-1),  3, PRIMARY),
        ]))
        story.append(c_block)
        for r in c.replies.filter(is_deleted=False).order_by('created_at'):
            r_block = Table([[
                Spacer(6*mm, 1),
                [Paragraph(f'{_safe(r.author.username)} (reply)',
                           S(f'ra{str(r.id)[:8]}', fontSize=8, textColor=GRAY, fontName='Helvetica-Bold')),
                 Paragraph(_safe(r.content), s_italic)],
            ]], colWidths=[6*mm, W - 6*mm])
            r_block.setStyle(TableStyle([
                ('TOPPADDING',    (0,0), (-1,-1), 3),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
                ('LEFTPADDING',   (0,0), (-1,-1), 0),
                ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ]))
            story.append(r_block)
        story.append(Spacer(1, 3*mm))

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{_filename(idea_id, idea.title, "pdf")}"'
    return resp


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_pdf(request, idea_id):
    idea, feedbacks, comments, err = _load_idea(idea_id, request.user)
    if err:
        return err

    from datetime import datetime
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        HRFlowable,
        KeepTogether,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    primary = colors.HexColor('#681A15')
    ink = colors.HexColor('#211716')
    muted = colors.HexColor('#756664')
    soft = colors.HexColor('#F5F0EF')
    softer = colors.HexColor('#FBF8F7')
    border = colors.HexColor('#E5D8D6')
    gold = colors.HexColor('#B98727')
    green = colors.HexColor('#397A4E')

    fb_list = list(feedbacks)
    comment_list = list(comments)
    replies_count = sum(c.replies.filter(is_deleted=False).count() for c in comment_list)
    comments_total = len(comment_list) + replies_count
    avg = _avg(fb_list)
    generated_at = datetime.utcnow().strftime('%B %d, %Y at %H:%M UTC')

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f'IdeaLab Report - {idea.title}',
        author='IdeaLab',
    )
    width = doc.width

    def st(name, **kwargs):
        return ParagraphStyle(name, **kwargs)

    title_style = st('report-title', fontName='Helvetica-Bold', fontSize=24, leading=29, textColor=colors.white)
    eyebrow_style = st('report-eyebrow', fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.HexColor('#F4DAD8'))
    cover_meta_style = st('report-cover-meta', fontName='Helvetica', fontSize=9, leading=13, textColor=colors.HexColor('#F7E7E5'))
    section_style = st('report-section', fontName='Helvetica-Bold', fontSize=10, leading=13, textColor=primary, spaceBefore=11, spaceAfter=7)
    body_style = st('report-body', fontName='Helvetica', fontSize=9.5, leading=14, textColor=ink, spaceAfter=4)
    small_style = st('report-small', fontName='Helvetica', fontSize=7.5, leading=10, textColor=muted)
    label_style = st('report-label', fontName='Helvetica-Bold', fontSize=7.5, leading=10, textColor=muted)
    metric_style = st('report-metric', fontName='Helvetica-Bold', fontSize=17, leading=20, textColor=primary, alignment=TA_CENTER)
    note_style = st('report-note', fontName='Helvetica-Oblique', fontSize=8.5, leading=12, textColor=colors.HexColor('#4A3E3C'))

    def para(text, paragraph_style=body_style):
        return Paragraph(_safe(str(text)) if text not in (None, '') else '-', paragraph_style)

    def section(title):
        return [
            HRFlowable(width='100%', thickness=0.7, color=border, spaceBefore=5, spaceAfter=5),
            Paragraph(title, section_style),
        ]

    def card(rows, col_widths=None, background=colors.white):
        table = Table(rows, colWidths=col_widths or [width])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), background),
            ('BOX', (0, 0), (-1, -1), 0.7, border),
            ('TOPPADDING', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 9),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        return table

    def metric(value, label, sublabel=''):
        items = [Paragraph(str(value), metric_style), Paragraph(label, label_style)]
        if sublabel:
            items.append(Paragraph(sublabel, small_style))
        return items

    def status_color():
        if idea.status == 'validated':
            return green
        if idea.status == 'review':
            return gold
        if idea.status == 'rejected':
            return primary
        return muted

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setStrokeColor(border)
        canvas.line(doc_obj.leftMargin, 12 * mm, A4[0] - doc_obj.rightMargin, 12 * mm)
        canvas.setFont('Helvetica-Bold', 8)
        canvas.setFillColor(primary)
        canvas.drawString(doc_obj.leftMargin, 8 * mm, 'IdeaLab')
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(muted)
        canvas.drawCentredString(A4[0] / 2, 8 * mm, f'Idea #{idea.id} - {idea.title[:42]}')
        canvas.drawRightString(A4[0] - doc_obj.rightMargin, 8 * mm, f'Page {doc_obj.page}')
        canvas.restoreState()

    story = []

    status_badge = Table([[Paragraph(idea.status.upper(), st(
        'status-badge', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white, alignment=TA_CENTER
    ))]], colWidths=[34 * mm])
    status_badge.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), status_color()),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    cover = Table([
        [Paragraph('IDEALAB IDEA REPORT', eyebrow_style), status_badge],
        [Paragraph(_safe(idea.title), title_style), ''],
        [Paragraph(f'Prepared for {_safe(idea.owner.username)} on {generated_at}', cover_meta_style), ''],
    ], colWidths=[width - 40 * mm, 40 * mm])
    cover.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), primary),
        ('SPAN', (0, 1), (1, 1)),
        ('SPAN', (0, 2), (1, 2)),
        ('TOPPADDING', (0, 0), (-1, -1), 14),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 16),
        ('RIGHTPADDING', (0, 0), (-1, -1), 16),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.extend([cover, Spacer(1, 7 * mm)])

    summary = Table([[
        metric(avg, 'Average SGV', 'out of 25'),
        metric(len(fb_list), 'Reviews', 'submitted'),
        metric(comments_total, 'Comments', 'including replies'),
        metric(idea.status.capitalize(), 'Status'),
    ]], colWidths=[width / 4] * 4)
    summary.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), soft),
        ('BOX', (0, 0), (-1, -1), 0.7, border),
        ('LINEAFTER', (0, 0), (2, 0), 0.5, border),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.extend([summary, Spacer(1, 5 * mm)])

    meta = Table([
        [para('Owner', label_style), para(idea.owner.username), para('Sector', label_style), para(idea.sector)],
        [para('Created', label_style), para(idea.created_at.strftime('%B %d, %Y')), para('Idea ID', label_style), para(idea.id)],
    ], colWidths=[22 * mm, width / 2 - 22 * mm, 22 * mm, width / 2 - 22 * mm])
    meta.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), softer),
        ('BOX', (0, 0), (-1, -1), 0.5, border),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, border),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(meta)

    story.extend(section('Idea Overview'))
    for label, value in [
        ('Description', idea.description),
        ('Problem', idea.problem),
        ('Solution', idea.solution),
        ('Target Audience', idea.target),
    ]:
        story.append(card([[Paragraph(label.upper(), label_style)], [para(value)]], background=softer))
        story.append(Spacer(1, 2 * mm))

    story.extend(section('Score Breakdown'))
    if fb_list:
        dimensions = [
            ('Market', round(sum(f.market_score for f in fb_list) / len(fb_list), 1)),
            ('Innovation', round(sum(f.innovation_score for f in fb_list) / len(fb_list), 1)),
            ('Feasibility', round(sum(f.feasibility_score for f in fb_list) / len(fb_list), 1)),
            ('ROI', round(sum(f.roi_score for f in fb_list) / len(fb_list), 1)),
        ]
        score_rows = [[para('Dimension', label_style), para('Average', label_style), para('Visual', label_style)]]
        bar_width = 72 * mm
        for name, value in dimensions:
            filled = max(1 * mm, bar_width * min(value / 25, 1))
            fill = Table([['']], colWidths=[filled], rowHeights=[5 * mm])
            fill.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), primary)]))
            bar = Table([[fill, '']], colWidths=[filled, bar_width - filled], rowHeights=[5 * mm])
            bar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#EDE3E1')),
                ('BOX', (0, 0), (-1, -1), 0.25, border),
            ]))
            score_rows.append([para(name), para(f'{value}/25'), bar])

        score_table = Table(score_rows, colWidths=[width * 0.28, width * 0.18, width * 0.54])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), soft),
            ('BOX', (0, 0), (-1, -1), 0.7, border),
            ('INNERGRID', (0, 0), (-1, -1), 0.3, border),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.extend([score_table, Spacer(1, 3 * mm)])
    else:
        story.append(card([[para('No reviews have been submitted yet.', small_style)]], background=softer))

    story.extend(section(f'Reviewer Feedback ({len(fb_list)})'))
    if not fb_list:
        story.append(card([[para('No feedback available yet.', small_style)]], background=softer))

    for index, feedback in enumerate(fb_list, start=1):
        helpful = 'Most helpful' if feedback.is_helpful else ''
        feedback_header = Table([[
            para(f'{index}. {feedback.reviewer.username}', st(
                f'feedback-reviewer-{index}', fontName='Helvetica-Bold', fontSize=10, leading=13, textColor=ink
            )),
            para(_level(feedback), label_style),
            para(helpful, st(
                f'feedback-helpful-{index}', fontName='Helvetica-Bold', fontSize=8, leading=10,
                textColor=green, alignment=TA_CENTER
            )),
            para(f'{round(feedback.weighted_score or 0, 2)}/25', st(
                f'feedback-score-{index}', fontName='Helvetica-Bold', fontSize=12, leading=14,
                textColor=primary, alignment=TA_RIGHT
            )),
        ]], colWidths=[width * 0.38, width * 0.18, width * 0.22, width * 0.22])
        feedback_header.setStyle(TableStyle([
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        feedback_scores = Table([[
            metric(f'{feedback.market_score}/25', 'Market'),
            metric(f'{feedback.innovation_score}/25', 'Innovation'),
            metric(f'{feedback.feasibility_score}/25', 'Feasibility'),
            metric(f'{feedback.roi_score}/25', 'ROI'),
        ]], colWidths=[width / 4] * 4)
        feedback_scores.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), soft),
            ('LINEAFTER', (0, 0), (2, 0), 0.5, border),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))

        block = card([
            [feedback_header],
            [feedback_scores],
            [Paragraph('COMMENT', label_style)],
            [para(feedback.comment, note_style)],
            [para(feedback.created_at.strftime('%B %d, %Y at %H:%M'), small_style)],
        ])
        story.append(KeepTogether([block, Spacer(1, 4 * mm)]))

    story.extend(section(f'Community Discussion ({comments_total})'))
    if not comment_list:
        story.append(card([[para('No comments yet.', small_style)]], background=softer))

    for comment_index, comment in enumerate(comment_list, start=1):
        comment_block = card([[
            [
                para(f'{comment_index}. {comment.author.username}', st(
                    f'comment-author-{comment_index}', fontName='Helvetica-Bold', fontSize=9,
                    leading=12, textColor=primary
                )),
                para(comment.content),
            ],
            para(comment.created_at.strftime('%b %d, %Y'), small_style),
        ]], col_widths=[width * 0.82, width * 0.18], background=softer)
        story.append(comment_block)

        for reply_index, reply in enumerate(comment.replies.filter(is_deleted=False).order_by('created_at'), start=1):
            reply_block = Table([[
                '',
                [
                    para(f'{reply.author.username} replied', st(
                        f'reply-author-{comment_index}-{reply_index}',
                        fontName='Helvetica-Bold',
                        fontSize=8,
                        leading=10,
                        textColor=muted,
                    )),
                    para(reply.content, note_style),
                ],
            ]], colWidths=[8 * mm, width - 8 * mm])
            reply_block.setStyle(TableStyle([
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(reply_block)
        story.append(Spacer(1, 3 * mm))

    story.extend(section('Export Notes'))
    story.append(card([[
        para(
            'This report includes the idea details, reviewer scoring, submitted feedback, and community discussion available at export time. Scores are displayed on the same 0-25 scale used by IdeaLab review dimensions.',
            small_style,
        )
    ]], background=softer))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{_filename(idea_id, idea.title, "pdf")}"'
    return resp
